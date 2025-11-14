import os
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes, ConversationHandler
from bot.utils.states import AdminStates

async def admin_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Главное меню администратора"""
    from bot.utils.config import ADMIN_CHAT_IDS

    user = update.effective_user

    if user.id not in ADMIN_CHAT_IDS:
        await update.message.reply_text("У вас нет прав администратора.")
        return ConversationHandler.END
    
    context.bot_data['admin_user_id'] = user.id
    
    keyboard = [
        [InlineKeyboardButton("📋 Текущие заявки", callback_data='view_requests')],
        [InlineKeyboardButton("📦 Архив заявок", callback_data='view_archive')],
        [InlineKeyboardButton("📤 Выгрузить в Excel", callback_data='export_excel')],
        [InlineKeyboardButton("🗑️ Очистить старые заявки", callback_data='cleanup_requests')],
        [InlineKeyboardButton("📝 Управление группами", callback_data='manage_groups')],
        [InlineKeyboardButton("🎯 Управление целями печати", callback_data='manage_purposes')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        "🔧 Панель администратора\n\nВыберите действие:",
        reply_markup=reply_markup
    )
    
    return AdminStates.VIEW_REQUESTS

async def view_requests(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Просмотр текущих заявок"""
    query = update.callback_query
    await query.answer()
    
    db = context.bot_data.get('db')
    
    if not db:
        await query.edit_message_text("Ошибка: база данных не инициализирована.")
        return ConversationHandler.END
    
    try:
        requests = db.get_all_requests()
        
        if not requests:
            keyboard = [[InlineKeyboardButton("🔙 Главное меню", callback_data='admin_main_menu')]]
            await query.edit_message_text(
                "📋 Заявок пока нет.",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return AdminStates.VIEW_REQUESTS
        
        page = context.user_data.get('page', 0)
        items_per_page = 5
        
        start_idx = page * items_per_page
        end_idx = start_idx + items_per_page
        page_requests = requests[start_idx:end_idx]
        
        text = "📋 Список заявок:\n\n"
        
        keyboard = []
        
        for req in page_requests:
            status_emoji = {
                'В очереди': '⚪',
                'В работе': '🟡',
                'Готово': '🟢'
            }.get(req.get('status', ''), '⚪')
            
            user_info = f"@{req.get('username', 'нет')}" if req.get('username') else "нет username"
            
            text += (
                f"{status_emoji} #{req.get('id')[:8]}\n"
                f"👤 {req.get('first_name')} {req.get('last_name')}\n"
                f"🔗 {user_info}\n"
                f"📚 {req.get('group')} | 🎯 {req.get('purpose')}\n"
                f"📅 {req.get('date')}\n"
                f"───────────────\n"
            )
            
            # Кнопка для детального просмотра
            keyboard.append([
                InlineKeyboardButton(
                    f"📄 Детали #{req.get('id')[:6]}",
                    callback_data=f"detail_{req.get('id')}"
                )
            ])
        
        # Навигация
        nav_buttons = []
        if page > 0:
            nav_buttons.append(InlineKeyboardButton("◀️ Назад", callback_data='prev_page'))
        if end_idx < len(requests):
            nav_buttons.append(InlineKeyboardButton("Вперед ▶️", callback_data='next_page'))
        
        if nav_buttons:
            keyboard.append(nav_buttons)
        
        keyboard.append([InlineKeyboardButton("🔙 Главное меню", callback_data='admin_main_menu')])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(text, reply_markup=reply_markup)
        
    except Exception as e:
        print(f"Ошибка при просмотре заявок: {e}")
        await query.edit_message_text(f"Произошла ошибка: {e}")
    
    return AdminStates.VIEW_REQUESTS

async def view_request_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Детальный просмотр заявки"""
    query = update.callback_query
    await query.answer()
    
    request_id = query.data.split('_')[1]
    db = context.bot_data.get('db')
    
    request_data = db.get_request_by_id(request_id)
    
    if not request_data:
        await query.answer("Заявка не найдена!")
        return AdminStates.VIEW_REQUESTS
    
    status_emoji = {
        'В очереди': '⚪',
        'В работе': '🟡',
        'Готово': '🟢'
    }.get(request_data.get('status', ''), '⚪')
    
    user_info = f"@{request_data.get('username', 'нет')}" if request_data.get('username') else "нет username"
    comment = request_data.get('comment', '')
    
    text = (
        f"📄 Детали заявки #{request_id[:8]}\n\n"
        f"{status_emoji} Статус: {request_data.get('status')}\n"
        f"👤 Имя: {request_data.get('first_name')} {request_data.get('last_name')}\n"
        f"🔗 Username: {user_info}\n"
        f"📚 Группа: {request_data.get('group')}\n"
        f"🎯 Цель печати: {request_data.get('purpose')}\n"
        f"📅 Дата подачи: {request_data.get('date')}\n"
        f"📎 Файл: {request_data.get('file_name', 'нет')}\n"
    )
    
    if comment:
        text += f"\n💬 Комментарий: {comment}\n"
    
    # Кнопки действий
    keyboard = []
    
    status = request_data.get('status')
    
    if status == 'В очереди':
        keyboard.append([InlineKeyboardButton("✅ Принять в работу", callback_data=f"accept_{request_id}")])
    elif status == 'В работе':
        keyboard.append([InlineKeyboardButton("✔️ Готово", callback_data=f"complete_{request_id}")])
    elif status == 'Готово':
        keyboard.append([InlineKeyboardButton("📦 В архив", callback_data=f"archive_{request_id}")])
    
    # Кнопки для всех статусов
    # СТАЛО:
    keyboard.append([InlineKeyboardButton("📥 Получить файл", callback_data=f"send_file_admin_{request_id}")])
    keyboard.append([InlineKeyboardButton("💬 Добавить комментарий", callback_data=f"add_comment_{request_id}")])
    
    if request_data.get('telegram_id'):
        keyboard.append([InlineKeyboardButton("✉️ Написать пользователю", callback_data=f"message_user_{request_id}")])
    
    keyboard.append([InlineKeyboardButton("🔙 К списку заявок", callback_data='view_requests')])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup)
    
    return AdminStates.VIEW_REQUESTS

async def accept_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Принять заявку в работу"""
    query = update.callback_query
    await query.answer()
    
    request_id = query.data.split('_')[1]
    db = context.bot_data.get('db')
    
    try:
        db.update_status(request_id, 'В работе')
        
        request_data = db.get_request_by_id(request_id)
        if request_data and request_data.get('telegram_id'):
            try:
                await context.bot.send_message(
                    chat_id=int(request_data['telegram_id']),
                    text=f"📢 Ваша заявка #{request_id[:8]} принята в работу!"
                )
            except Exception as e:
                print(f"Не удалось отправить уведомление пользователю: {e}")
        
        await query.answer("✅ Заявка принята в работу!")
        
        # Обновляем детали заявки
        context.user_data['temp_callback_data'] = f"detail_{request_id}"
        await view_request_detail(update, context)
        
    except Exception as e:
        print(f"Ошибка при принятии заявки: {e}")
        await query.answer(f"Ошибка: {e}")
    
    return AdminStates.VIEW_REQUESTS

async def complete_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отметить заявку как готовую"""
    query = update.callback_query
    await query.answer()
    
    request_id = query.data.split('_')[1]
    db = context.bot_data.get('db')
    
    try:
        db.update_status(request_id, 'Готово')
        
        request_data = db.get_request_by_id(request_id)
        if request_data and request_data.get('telegram_id'):
            try:
                comment = request_data.get('comment', '')
                message = f"✅ Ваша заявка #{request_id[:8]} готова!"
                if comment:
                    message += f"\n\n💬 {comment}"
                
                await context.bot.send_message(
                    chat_id=int(request_data['telegram_id']),
                    text=message
                )
            except Exception as e:
                print(f"Не удалось отправить уведомление пользователю: {e}")
        
        await query.answer("✅ Заявка готова!")
        
        # Обновляем детали заявки
        context.user_data['temp_callback_data'] = f"detail_{request_id}"
        await view_request_detail(update, context)
        
    except Exception as e:
        print(f"Ошибка при завершении заявки: {e}")
        await query.answer(f"Ошибка: {e}")
    
    return AdminStates.VIEW_REQUESTS

async def archive_request(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Переместить заявку в архив"""
    query = update.callback_query
    await query.answer()
    
    request_id = query.data.split('_')[1]
    db = context.bot_data.get('db')
    
    try:
        if db.archive_request(request_id):
            await query.answer("📦 Заявка перемещена в архив!")
            await view_requests(update, context)
        else:
            await query.answer("Ошибка при архивации!")
        
    except Exception as e:
        print(f"Ошибка при архивации заявки: {e}")
        await query.answer(f"Ошибка: {e}")
    
    return AdminStates.VIEW_REQUESTS

async def send_file_to_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправить файл администратору (тому, кто нажал кнопку)"""
    query = update.callback_query
    await query.answer()

    # Извлекаем request_id из callback_data вида 'send_file_admin_<id>'
    try:
        request_id = query.data.split('_', 3)[3]
    except IndexError:
        await query.answer("Ошибка: некорректный ID заявки.")
        return AdminStates.VIEW_REQUESTS

    db = context.bot_data.get('db')
    if not db:
        await query.answer("Ошибка: база данных недоступна.")
        return AdminStates.VIEW_REQUESTS

    request_data = db.get_request_by_id(request_id)
    if not request_data:
        await query.answer("Заявка не найдена!")
        return AdminStates.VIEW_REQUESTS

    file_path = request_data.get('file_path')
    if not file_path or not os.path.exists(file_path):
        await query.answer("Файл не найден на сервере!")
        return AdminStates.VIEW_REQUESTS

    admin_chat_id = update.effective_user.id  # Тот, кто нажал кнопку

    try:
        with open(file_path, 'rb') as f:
            await context.bot.send_document(
                chat_id=admin_chat_id,
                document=f,
                caption=(
                    f"📎 Файл заявки #{request_id[:8]}\n"
                    f"👤 {request_data.get('first_name', '')} {request_data.get('last_name', '')}\n"
                    f"📚 Группа: {request_data.get('group', '—')}\n"
                    f"🎯 Цель: {request_data.get('purpose', '—')}"
                )
            )
        await query.answer("✅ Файл отправлен вам!")
    except Exception as e:
        print(f"Ошибка при отправке файла админу: {e}")
        await query.answer(f"❌ Ошибка: {e}")

    return AdminStates.VIEW_REQUESTS

async def start_add_comment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начать добавление комментария"""
    query = update.callback_query
    await query.answer()
    
    request_id = query.data.split('_', 2)[2]
    context.user_data['comment_request_id'] = request_id
    
    await query.edit_message_text(
        "💬 Введите комментарий для пользователя (например, где забрать изделие):"
    )
    
    return AdminStates.ADDING_COMMENT

async def save_comment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Сохранить комментарий"""
    comment = update.message.text
    request_id = context.user_data.get('comment_request_id')
    
    if not request_id:
        await update.message.reply_text("Ошибка: заявка не найдена.")
        return ConversationHandler.END
    
    db = context.bot_data.get('db')
    
    if db.add_comment(request_id, comment):
        await update.message.reply_text(f"✅ Комментарий добавлен!\n\n💬 {comment}")
        
        # Показываем кнопку возврата
        keyboard = [[InlineKeyboardButton("🔙 К деталям заявки", callback_data=f"detail_{request_id}")]]
        await update.message.reply_text(
            "Выберите действие:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        await update.message.reply_text("Ошибка при добавлении комментария.")
    
    return AdminStates.VIEW_REQUESTS

async def start_message_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начать написание сообщения пользователю"""
    query = update.callback_query
    await query.answer()
    
    request_id = query.data.split('_', 2)[2]
    context.user_data['message_request_id'] = request_id
    
    db = context.bot_data.get('db')
    request_data = db.get_request_by_id(request_id)
    
    user_name = f"{request_data.get('first_name')} {request_data.get('last_name')}"
    
    await query.edit_message_text(
        f"✉️ Напишите сообщение для пользователя {user_name}:"
    )
    
    return AdminStates.MESSAGING_USER

async def send_message_to_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправить сообщение пользователю"""
    message_text = update.message.text
    request_id = context.user_data.get('message_request_id')
    
    if not request_id:
        await update.message.reply_text("Ошибка: заявка не найдена.")
        return ConversationHandler.END
    
    db = context.bot_data.get('db')
    request_data = db.get_request_by_id(request_id)
    
    if not request_data or not request_data.get('telegram_id'):
        await update.message.reply_text("Ошибка: не удалось найти пользователя.")
        return ConversationHandler.END
    
    try:
        await context.bot.send_message(
            chat_id=int(request_data['telegram_id']),
            text=f"💬 Сообщение от администратора:\n\n{message_text}\n\n📝 Касательно заявки #{request_id[:8]}"
        )
        
        await update.message.reply_text("✅ Сообщение отправлено!")
        
        keyboard = [[InlineKeyboardButton("🔙 К деталям заявки", callback_data=f"detail_{request_id}")]]
        await update.message.reply_text(
            "Выберите действие:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
    except Exception as e:
        print(f"Ошибка при отправке сообщения: {e}")
        await update.message.reply_text(f"Ошибка: {e}")
    
    return AdminStates.VIEW_REQUESTS

async def view_archive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Просмотр архива"""
    query = update.callback_query
    await query.answer()
    
    db = context.bot_data.get('db')
    archive = db.get_archive()
    
    if not archive:
        keyboard = [[InlineKeyboardButton("🔙 Главное меню", callback_data='admin_main_menu')]]
        await query.edit_message_text(
            "📦 Архив пуст.",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return AdminStates.VIEW_REQUESTS
    
    text = f"📦 Архив заявок ({len(archive)} шт.):\n\n"
    
    for req in archive[-10:]:  # Последние 10 заявок
        text += (
            f"#{req.get('id')[:8]} | {req.get('first_name')} {req.get('last_name')}\n"
            f"📅 {req.get('archived_date', 'н/д')}\n"
            f"───────────────\n"
        )
    
    if len(archive) > 10:
        text += f"\n... и еще {len(archive) - 10} заявок"
    
    keyboard = [
        [InlineKeyboardButton("🔙 Главное меню", callback_data='admin_main_menu')]
    ]
    
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    
    return AdminStates.VIEW_REQUESTS

async def cleanup_requests(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ручная очистка старых заявок"""
    query = update.callback_query
    await query.answer()
    
    db = context.bot_data.get('db')
    
    result = db.manual_cleanup()
    
    text = (
        f"🗑️ Очистка завершена!\n\n"
        f"📦 Перемещено в архив: {result['moved_to_archive']}\n"
        f"🗑️ Удалено из архива (старше 2 недель): {result['cleaned_from_archive']}"
    )
    
    keyboard = [[InlineKeyboardButton("🔙 Главное меню", callback_data='admin_main_menu')]]
    
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    
    return AdminStates.VIEW_REQUESTS

async def navigate_pages(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Навигация по страницам"""
    query = update.callback_query
    await query.answer()
    
    if query.data == 'next_page':
        context.user_data['page'] = context.user_data.get('page', 0) + 1
    elif query.data == 'prev_page':
        context.user_data['page'] = max(0, context.user_data.get('page', 0) - 1)
    
    await view_requests(update, context)
    
    return AdminStates.VIEW_REQUESTS

async def back_to_admin_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Вернуться в главное меню админа"""
    query = update.callback_query
    await query.answer()
    
    context.user_data['page'] = 0
    
    keyboard = [
        [InlineKeyboardButton("📋 Текущие заявки", callback_data='view_requests')],
        [InlineKeyboardButton("📦 Архив заявок", callback_data='view_archive')],
        [InlineKeyboardButton("🗑️ Очистить старые заявки", callback_data='cleanup_requests')],
        [InlineKeyboardButton("📝 Управление группами", callback_data='manage_groups')],
        [InlineKeyboardButton("🎯 Управление целями печати", callback_data='manage_purposes')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        "🔧 Панель администратора\n\nВыберите действие:",
        reply_markup=reply_markup
    )
    
    return AdminStates.VIEW_REQUESTS

# Управление группами и целями (оставляем как было)

async def manage_groups(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    groups = context.bot_data.get('groups', [])
    
    text = "📝 Управление группами:\n\n"
    text += "Текущие группы:\n"
    
    if groups:
        for idx, group in enumerate(groups, 1):
            text += f"{idx}. {group}\n"
    else:
        text += "Список пуст\n"
    
    keyboard = [
        [InlineKeyboardButton("➕ Добавить группу", callback_data='add_group')],
        [InlineKeyboardButton("➖ Удалить группу", callback_data='remove_group')],
        [InlineKeyboardButton("🔙 Главное меню", callback_data='admin_main_menu')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup)
    
    return AdminStates.MANAGE_GROUPS

async def manage_purposes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    purposes = context.bot_data.get('purposes', [])
    
    text = "🎯 Управление целями печати:\n\n"
    text += "Текущие цели:\n"
    
    if purposes:
        for idx, purpose in enumerate(purposes, 1):
            text += f"{idx}. {purpose}\n"
    else:
        text += "Список пуст\n"
    
    keyboard = [
        [InlineKeyboardButton("➕ Добавить цель", callback_data='add_purpose')],
        [InlineKeyboardButton("➖ Удалить цель", callback_data='remove_purpose')],
        [InlineKeyboardButton("🔙 Главное меню", callback_data='admin_main_menu')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup)
    
    return AdminStates.MANAGE_PURPOSES

async def add_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    await query.edit_message_text("Введите название новой группы:")
    
    return AdminStates.ADDING_GROUP

async def save_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    group_name = update.message.text
    groups = context.bot_data.get('groups', [])
    
    if group_name not in groups:
        groups.append(group_name)
        context.bot_data['groups'] = groups
        await update.message.reply_text(f"✅ Группа '{group_name}' добавлена!")
    else:
        await update.message.reply_text(f"Группа '{group_name}' уже существует.")
    
    keyboard = [[InlineKeyboardButton("🔙 К управлению группами", callback_data='manage_groups')]]
    await update.message.reply_text("Выберите действие:", reply_markup=InlineKeyboardMarkup(keyboard))
    
    return AdminStates.MANAGE_GROUPS

async def remove_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    groups = context.bot_data.get('groups', [])
    
    if not groups:
        await query.answer("Список групп пуст!")
        return AdminStates.MANAGE_GROUPS
    
    keyboard = []
    for group in groups:
        keyboard.append([InlineKeyboardButton(f"❌ {group}", callback_data=f"delete_group_{group}")])
    
    keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data='manage_groups')])
    
    await query.edit_message_text(
        "Выберите группу для удаления:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    
    return AdminStates.MANAGE_GROUPS

async def delete_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    group_name = query.data.split('delete_group_')[1]
    groups = context.bot_data.get('groups', [])
    
    if group_name in groups:
        groups.remove(group_name)
        context.bot_data['groups'] = groups
        await query.answer(f"Группа '{group_name}' удалена!")
    
    await manage_groups(update, context)
    
    return AdminStates.MANAGE_GROUPS

async def add_purpose(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    await query.edit_message_text("Введите название новой цели печати:")
    
    return AdminStates.ADDING_PURPOSE

async def save_purpose(update: Update, context: ContextTypes.DEFAULT_TYPE):
    purpose_name = update.message.text
    purposes = context.bot_data.get('purposes', [])
    
    if purpose_name not in purposes:
        purposes.append(purpose_name)
        context.bot_data['purposes'] = purposes
        await update.message.reply_text(f"✅ Цель '{purpose_name}' добавлена!")
    else:
        await update.message.reply_text(f"Цель '{purpose_name}' уже существует.")
    
    keyboard = [[InlineKeyboardButton("🔙 К управлению целями", callback_data='manage_purposes')]]
    await update.message.reply_text("Выберите действие:", reply_markup=InlineKeyboardMarkup(keyboard))
    
    return AdminStates.MANAGE_PURPOSES

async def remove_purpose(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    purposes = context.bot_data.get('purposes', [])
    
    if not purposes:
        await query.answer("Список целей пуст!")
        return AdminStates.MANAGE_PURPOSES
    
    keyboard = []
    for purpose in purposes:
        keyboard.append([InlineKeyboardButton(f"❌ {purpose}", callback_data=f"delete_purpose_{purpose}")])
    
    keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data='manage_purposes')])
    
    await query.edit_message_text(
        "Выберите цель для удаления:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    
    return AdminStates.MANAGE_PURPOSES

async def delete_purpose(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    purpose_name = query.data.split('delete_purpose_')[1]
    purposes = context.bot_data.get('purposes', [])
    
    if purpose_name in purposes:
        purposes.remove(purpose_name)
        context.bot_data['purposes'] = purposes
        await query.answer(f"Цель '{purpose_name}' удалена!")
    
    await manage_purposes(update, context)
    
    return AdminStates.MANAGE_PURPOSES
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

async def export_to_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Выгрузить заявки в Excel с цветовой индикацией статусов"""
    query = update.callback_query
    await query.answer()

    db = context.bot_data.get('db')
    if not db:
        await query.edit_message_text("❌ База данных не доступна.")
        return AdminStates.VIEW_REQUESTS

    try:
        active = db.get_all_requests() or []
        archive = db.get_archive() or []
        all_requests = active + archive

        if not all_requests:
            await query.edit_message_text("📭 Нет заявок для выгрузки.")
            return AdminStates.VIEW_REQUESTS

        # Подготовка данных
        rows = []
        for req in all_requests:
            rows.append({
                "ID": req.get('id', '')[:8],
                "Статус": req.get('status', ''),
                "Имя": req.get('first_name', ''),
                "Фамилия": req.get('last_name', ''),
                "Группа": req.get('group', ''),
                "Цель": req.get('purpose', ''),
                "Файл": req.get('file_name', ''),
                "Комментарий": req.get('comment', ''),
                "Дата подачи": req.get('date', ''),
                "Telegram ID": req.get('telegram_id', ''),
                "Username": f"@{req.get('username')}" if req.get('username') else "",
            })

        df = pd.DataFrame(rows)

        # Путь к файлу
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"3d_print_requests_{timestamp}.xlsx"
        filepath = os.path.join("data", filename)
        os.makedirs("data", exist_ok=True)

        # Создаём Excel вручную через openpyxl (чтобы добавить цвета)
        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки"

        # Записываем заголовки
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)

        # Записываем данные
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Цвета для статусов
        status_colors = {
            "В очереди": "FFFFCC",   # светло-жёлтый
            "В работе": "FFFF00",    # жёлтый
            "Готово": "CCFFCC",      # светло-зелёный
            "Архив": "DDDDDD",       # серый
        }

        # Применяем цвета к строкам по статусу (столбец B = "Статус")
        for row in range(2, len(df) + 2):
            status_cell = ws[f"B{row}"]
            status = status_cell.value
            if status in status_colors:
                fill = PatternFill(start_color=status_colors[status], end_color=status_colors[status], fill_type="solid")
                for col in range(1, len(df.columns) + 1):
                    ws.cell(row=row, column=col).fill = fill

        # Сохраняем
        wb.save(filepath)

        # Отправляем админу
        admin_id = update.effective_user.id
        with open(filepath, 'rb') as f:
            await context.bot.send_document(
                chat_id=admin_id,
                document=f,
                caption="📊 Выгрузка заявок с цветовой индикацией статусов"
            )

        os.remove(filepath)
        await back_to_admin_menu(update, context)

    except Exception as e:
        print(f"Ошибка выгрузки Excel: {e}")
        await query.edit_message_text(f"❌ Ошибка: {e}")
        return AdminStates.VIEW_REQUESTS

    return AdminStates.VIEW_REQUESTS
